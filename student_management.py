from openpyxl import Workbook, load_workbook
import os


# Excel file will be saved in the same location as this Python file
FILE_NAME = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                         "student_data.xlsx")


# Create Excel file
def create_file():
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"

        ws.append([
            "Student Name",
            "Class",
            "Address",
            "Contact Number"
        ])

        wb.save(FILE_NAME)


# New Student
def new_student():
    print("\nNEW STUDENT")

    name = input("Enter student name: ").strip()
    student_class = input("Enter class: ").strip()
    address = input("Enter address: ").strip()
    contact = input("Enter contact number: ").strip()

    if name == "" or student_class == "" or address == "" or contact == "":
        print("All fields are required.")
        return

    if not contact.isdigit():
        print("Contact number must contain only digits.")
        return

    wb = load_workbook(FILE_NAME)
    ws = wb["Students"]

    ws.append([
        name,
        student_class,
        address,
        contact
    ])

    wb.save(FILE_NAME)

    print("Student added successfully.")


# View Student
def view_student():
    wb = load_workbook(FILE_NAME)
    ws = wb["Students"]

    if ws.max_row == 1:
        print("\nNo student records found.")
        return

    print("\nSTUDENT RECORDS\n")

    for row in ws.iter_rows(min_row=2, values_only=True):
        print("Student Name   :", row[0])
        print("Class          :", row[1])
        print("Address        :", row[2])
        print("Contact Number :", row[3])
        print()


# Delete Student
def delete_student():
    name = input("\nEnter student name to delete: ").strip()

    wb = load_workbook(FILE_NAME)
    ws = wb["Students"]

    found = False

    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row, 1).value).lower() == name.lower():
            ws.delete_rows(row, 1)
            found = True
            break

    if found:
        wb.save(FILE_NAME)
        print("Student deleted successfully.")
    else:
        print("Student not found.")


# Menu
def menu():
    while True:
        print("\n===== STUDENT MANAGEMENT SYSTEM =====")
        print("1. New Student")
        print("2. View Student")
        print("3. Delete Student")
        print("4. Exit")

        choice = input("Enter your choice: ").strip()

        if choice == "1":
            new_student()

        elif choice == "2":
            view_student()

        elif choice == "3":
            delete_student()

        elif choice == "4":
            print("Thank you.")
            break

        else:
            print("Invalid choice. Please enter 1, 2, 3 or 4.")


# Start program
create_file()
menu()
